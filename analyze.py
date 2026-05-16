import openpyxl
from rapidfuzz import fuzz, process
import unicodedata

def normalize(s):
    if not s or not isinstance(s, str):
        return ''
    s = s.strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s

wb1 = openpyxl.load_workbook('Points lista.xlsx')
ws1 = wb1.active
wb2 = openpyxl.load_workbook('clientes-zona-actualizado.xlsx')
ws2 = wb2.active

# Points: list of (nombre_norm, wkt, nombre_original)
points = []
for row in ws1.iter_rows(min_row=2, values_only=True):
    if row[0] and row[1] and isinstance(row[1], str):
        points.append((normalize(row[1]), row[0], row[1]))

# Clientes sin WKT: (row_num, norm_razon, orig_razon, codigo)
sin_wkt = []
for i, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
    if row[0] is None and row[3] and isinstance(row[3], str):
        sin_wkt.append((i, normalize(row[3]), row[3], row[2]))

points_names = [p[0] for p in points]

thresholds = [90, 80, 70, 60]
results = {}
for t in thresholds:
    results[t] = {'auto': 0, 'conflict': 0, 'nomatch': 0}

sample = sin_wkt[:200]

for row_num, norm_razon, orig_razon, codigo in sample:
    matches = process.extract(norm_razon, points_names, scorer=fuzz.token_sort_ratio, limit=3)
    for t in thresholds:
        above = [m for m in matches if m[1] >= t]
        if len(above) == 0:
            results[t]['nomatch'] += 1
        elif len(above) == 1:
            results[t]['auto'] += 1
        else:
            results[t]['conflict'] += 1

print("Umbral | Auto-fill | Conflicto | Sin match | (sobre 200 muestras)")
for t in thresholds:
    r = results[t]
    print(f"  {t}%  |   {r['auto']:4d}    |   {r['conflict']:4d}    |   {r['nomatch']:4d}")

print()
print("=== Ejemplos matches al 70% ===")
count = 0
for row_num, norm_razon, orig_razon, codigo in sample[:80]:
    matches = process.extract(norm_razon, points_names, scorer=fuzz.token_sort_ratio, limit=3)
    above = [m for m in matches if m[1] >= 70]
    if above and count < 15:
        point_matches = [(points[m[2]][2], m[1]) for m in above]
        print(f"  RAZON: {repr(orig_razon):40s} -> {point_matches}")
        count += 1

print()
print("=== Ejemplos SIN match al 70% ===")
count = 0
for row_num, norm_razon, orig_razon, codigo in sample:
    matches = process.extract(norm_razon, points_names, scorer=fuzz.token_sort_ratio, limit=3)
    above = [m for m in matches if m[1] >= 70]
    if not above and count < 10:
        best = matches[0] if matches else None
        print(f"  RAZON: {repr(orig_razon):40s}  best_score={best[1] if best else 'N/A'} ({repr(points[best[2]][2]) if best else ''})")
        count += 1
