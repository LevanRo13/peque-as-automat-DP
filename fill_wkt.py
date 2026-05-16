import openpyxl
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz, process
import unicodedata
import shutil

# ── Colores ────────────────────────────────────────────────────────────────
FILL_AUTO   = PatternFill("solid", fgColor="C6EFCE")  # verde   >= 90%, auto
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # amarillo 70-89%, unico
FILL_ORANGE = PatternFill("solid", fgColor="FFCC66")  # naranja  70-89%, conflicto
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")  # rojo     <70%, sin match

# ── Config ─────────────────────────────────────────────────────────────────
INPUT_FILE  = "clientes-zona-actualizado.xlsx"
POINTS_FILE = "Points lista.xlsx"
OUTPUT_FILE = "clientes-zona-resultado.xlsx"

THRESHOLD_AUTO  = 90  # >= este score -> auto-fill verde
THRESHOLD_FUZZY = 70  # >= este score -> revision amarillo/naranja
TOP_N           = 3   # cuantos candidatos mostrar en col L para conflictos

# ── Helpers ────────────────────────────────────────────────────────────────
def normalize(s):
    if not s or not isinstance(s, str):
        return ''
    s = s.strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s

# ── Cargar Points lista ────────────────────────────────────────────────────
wb_pts = openpyxl.load_workbook(POINTS_FILE)
ws_pts = wb_pts.active

# pool: lista de [nombre_norm, wkt, nombre_original, usado]
# usamos lista con indice para poder marcar como usado
pool = []
for row in ws_pts.iter_rows(min_row=2, values_only=True):
    wkt, nombre = row[0], row[1]
    if wkt and nombre and isinstance(nombre, str):
        pool.append([normalize(nombre), wkt, nombre, False])  # False = disponible

print(f"Points cargados: {len(pool)}")

# ── Copiar archivo destino ─────────────────────────────────────────────────
shutil.copy(INPUT_FILE, OUTPUT_FILE)
wb = openpyxl.load_workbook(OUTPUT_FILE)
ws = wb.active

# ── Marcar como usados los WKTs que ya estan en clientes-zona ─────────────
wkts_ya_usados = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        wkts_ya_usados.add(row[0])

for entry in pool:
    if entry[1] in wkts_ya_usados:
        entry[3] = True  # marcar como usado

usados_previos = sum(1 for e in pool if e[3])
print(f"Points ya usados por clientes existentes: {usados_previos}")
print(f"Points disponibles para asignar: {len(pool) - usados_previos}")

# ── Agregar columna de notas (col L) ──────────────────────────────────────
NOTES_COL = 12
if ws.cell(1, NOTES_COL).value is None:
    ws.cell(1, NOTES_COL).value = "CANDIDATOS_FUZZY"

# ── Counters ───────────────────────────────────────────────────────────────
count_already = 0
count_auto    = 0
count_yellow  = 0
count_orange  = 0
count_red     = 0

# ── Procesar cada fila ─────────────────────────────────────────────────────
def get_available_pool():
    """Retorna indices y nombres solo de los points NO usados."""
    return [(i, entry) for i, entry in enumerate(pool) if not entry[3]]

total_rows = ws.max_row
for row_idx in range(2, total_rows + 1):
    wkt_cell    = ws.cell(row_idx, 1)
    nombre_cell = ws.cell(row_idx, 2)
    razon_cell  = ws.cell(row_idx, 4)
    notes_cell  = ws.cell(row_idx, NOTES_COL)

    # Ya tiene WKT -> saltar
    if wkt_cell.value is not None:
        count_already += 1
        continue

    razon = razon_cell.value
    if not razon or not isinstance(razon, str):
        count_red += 1
        wkt_cell.fill    = FILL_RED
        nombre_cell.fill = FILL_RED
        continue

    norm_razon = normalize(razon)

    # Construir pool disponible en este momento
    available = get_available_pool()
    if not available:
        # No quedan points libres
        wkt_cell.fill    = FILL_RED
        nombre_cell.fill = FILL_RED
        notes_cell.value = "SIN MATCH - pool agotado"
        count_red += 1
        continue

    avail_indices = [i for i, _ in available]
    avail_names   = [entry[0] for _, entry in available]

    raw_matches = process.extract(
        norm_razon, avail_names,
        scorer=fuzz.token_sort_ratio,
        limit=TOP_N
    )

    # raw_matches[x][2] es el indice dentro de avail_names, lo mapeamos al pool real
    def get_pool_entry(match):
        return pool[avail_indices[match[2]]]

    above_auto  = [m for m in raw_matches if m[1] >= THRESHOLD_AUTO]
    above_fuzzy = [m for m in raw_matches if m[1] >= THRESHOLD_FUZZY]

    if above_auto:
        # Verde: auto-fill, marcar como usado
        best = above_auto[0]
        pt = get_pool_entry(best)
        pool[avail_indices[best[2]]][3] = True  # marcar usado
        wkt_cell.value    = pt[1]
        nombre_cell.value = pt[2]
        wkt_cell.fill     = FILL_AUTO
        nombre_cell.fill  = FILL_AUTO
        notes_cell.value  = f"AUTO {best[1]:.0f}% | {pt[2]}"
        count_auto += 1

    elif len(above_fuzzy) == 1:
        # Amarillo: unico candidato, marcar como usado
        best = above_fuzzy[0]
        pt = get_pool_entry(best)
        pool[avail_indices[best[2]]][3] = True  # marcar usado
        wkt_cell.value    = pt[1]
        nombre_cell.value = pt[2]
        wkt_cell.fill     = FILL_YELLOW
        nombre_cell.fill  = FILL_YELLOW
        notes_cell.value  = f"REVISAR {best[1]:.0f}% | {pt[2]}"
        count_yellow += 1

    elif len(above_fuzzy) > 1:
        # Naranja: conflicto — ponemos el mejor y lo marcamos como usado
        # La revision humana puede reasignarlo, pero evitamos duplicados
        best = above_fuzzy[0]
        pt = get_pool_entry(best)
        pool[avail_indices[best[2]]][3] = True  # marcar usado
        wkt_cell.value    = pt[1]
        nombre_cell.value = pt[2]
        wkt_cell.fill     = FILL_ORANGE
        nombre_cell.fill  = FILL_ORANGE
        candidatos = " | ".join(
            f"{get_pool_entry(m)[2]} ({m[1]:.0f}%)" for m in above_fuzzy
        )
        notes_cell.value = f"CONFLICTO | {candidatos}"
        count_orange += 1

    else:
        # Rojo: sin match
        wkt_cell.fill    = FILL_RED
        nombre_cell.fill = FILL_RED
        best = raw_matches[0] if raw_matches else None
        if best:
            pt = get_pool_entry(best)
            notes_cell.value = f"SIN MATCH | mejor: {pt[2]} ({best[1]:.0f}%)"
        else:
            notes_cell.value = "SIN MATCH"
        count_red += 1

wb.save(OUTPUT_FILE)

# ── Resumen ────────────────────────────────────────────────────────────────
total_sin_wkt = count_auto + count_yellow + count_orange + count_red
print()
print("=" * 55)
print(f"  Archivo generado: {OUTPUT_FILE}")
print("=" * 55)
print(f"  Ya tenian WKT (sin tocar):          {count_already:4d}")
print(f"  [VERDE]    Auto-fill (>=90%):        {count_auto:4d}")
print(f"  [AMARILLO] Revisar (70-89%, 1):      {count_yellow:4d}")
print(f"  [NARANJA]  Conflicto (70-89%, N):    {count_orange:4d}")
print(f"  [ROJO]     Sin match (<70%):         {count_red:4d}")
print(f"  -----------------------------------------")
print(f"  Total procesados sin WKT:            {total_sin_wkt:4d}")
print("=" * 55)
